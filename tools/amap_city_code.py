from datas import Datas
import os
from openpyxl import load_workbook

data = Datas('amap_city_code.json')

# 读取指定目录的excel文件
wb = load_workbook('.data/Amap_adcode_citycode.xlsx')
ws = wb.active

values = {}
for row in ws.iter_rows(min_row=2):  # 跳过表头
    location = row[0].value
    adcode = row[1].value
    citycode = row[2].value
    if(location is None or adcode is None or citycode is None):
        continue
    if citycode == '\\N':
        continue
    values[location] = {"adcode": adcode, "citycode": citycode}

data.store(values)
